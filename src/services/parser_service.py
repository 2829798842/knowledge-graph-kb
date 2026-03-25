"""从支持的导入文件类型中提取纯文本。"""

from dataclasses import dataclass
import re
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from docx import Document as DocxDocument
from docx.document import Document as DocxDocumentType
from docx.table import Table
from docx.text.paragraph import Paragraph
from pypdf import PdfReader

SUPPORTED_EXTENSIONS: tuple[str, ...] = (
    ".txt",
    ".pdf",
    ".docx",
    ".xlsx",
    ".xlsm",
    ".xls",
)
SUPPORTED_EXTENSION_SET: set[str] = set(SUPPORTED_EXTENSIONS)
SUPPORTED_EXTENSION_DISPLAY: str = "、".join(SUPPORTED_EXTENSIONS)

TEXT_EXTENSIONS: set[str] = {".txt"}
WORD_EXTENSIONS: set[str] = {".docx"}
OPENXML_SPREADSHEET_EXTENSIONS: set[str] = {".xlsx", ".xlsm"}
LEGACY_SPREADSHEET_EXTENSIONS: set[str] = {".xls"}

LIST_PREFIX_PATTERN = re.compile(r"^([*\-\u2022]|(\d+[\.\)]|[一二三四五六七八九十]+[、.)]))\s+")
SENTENCE_ENDING_PATTERN = re.compile(r"[。！？?!:：；;]$")
HEADER_VALUE_PATTERN = re.compile(r"^[A-Za-z_\-\u4e00-\u9fff][A-Za-z0-9_\-\s\u4e00-\u9fff/()%]*$")


class UnsupportedFileTypeError(ValueError):
    """当导入流水线不支持文件扩展名时抛出。"""


@dataclass(frozen=True)
class SpreadsheetWorksheetData:
    """表示从单个工作表中提取出的结构化行数据。"""

    name: str
    index: int
    rows: list[list[str]]


@dataclass(frozen=True)
class SpreadsheetDocumentData:
    """表示从单个工作簿文件中提取出的结构化表格数据。"""

    file_type: str
    worksheets: list[SpreadsheetWorksheetData]

    def to_text(self) -> str:
        return "\n\n".join(
            _format_sheet_block(worksheet.name, worksheet.rows)
            for worksheet in self.worksheets
            if worksheet.rows
        ).strip()


def detect_file_type(path: Path) -> str:
    """根据路径后缀识别文件类型。"""

    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSION_SET:
        display_suffix = suffix or "无扩展名"
        raise UnsupportedFileTypeError(
            f"暂不支持 {display_suffix}，目前仅支持 {SUPPORTED_EXTENSION_DISPLAY}"
        )
    return suffix.removeprefix(".")


def extract_text(path: Path) -> str:
    """从单个受支持文件中提取清洗后的文本。"""

    suffix = path.suffix.lower()
    if suffix in TEXT_EXTENSIONS:
        return _extract_text_file(path)
    if suffix == ".pdf":
        return _extract_pdf_text(path)
    if suffix in WORD_EXTENSIONS:
        return _extract_docx_text(path)
    if suffix in OPENXML_SPREADSHEET_EXTENSIONS | LEGACY_SPREADSHEET_EXTENSIONS:
        return _extract_spreadsheet_text(path)

    display_suffix = suffix or "无扩展名"
    raise UnsupportedFileTypeError(f"暂不支持 {display_suffix}，目前仅支持 {SUPPORTED_EXTENSION_DISPLAY}")


def _extract_text_file(path: Path) -> str:
    raw_bytes = path.read_bytes()
    candidate_encodings = _text_file_candidate_encodings(raw_bytes)
    raw_text: str | None = None

    for encoding in candidate_encodings:
        try:
            raw_text = raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
        if raw_text:
            break

    if raw_text is None:
        raw_text = raw_bytes.decode(encoding="utf-8", errors="ignore")

    cleaned_text = _cleanup_extracted_text(raw_text)
    if cleaned_text:
        return cleaned_text
    raise ValueError("文本文件中未提取到可用文本，请确认文件内容不是空白。")


def _extract_pdf_text(path: Path) -> str:
    try:
        pdf_reader = PdfReader(str(path))
        page_texts = [page.extract_text() or "" for page in pdf_reader.pages]
    except Exception as exc:  # noqa: BLE001
        raise ValueError("PDF 解析失败，请确认文件未损坏且内容可读取。") from exc

    cleaned_text = _cleanup_extracted_text("\n\n".join(page_texts))
    if cleaned_text:
        return cleaned_text
    raise ValueError("PDF 中未提取到可用文本，请确认文件不是扫描件或图片版。")


def _extract_docx_text(path: Path) -> str:
    try:
        document = DocxDocument(str(path))
    except Exception as exc:  # noqa: BLE001
        raise ValueError("Word 文档解析失败，请确认文件未损坏且为 .docx 格式。") from exc

    block_texts = _extract_docx_blocks(document)
    cleaned_text = _cleanup_extracted_text("\n\n".join(block_texts))
    if cleaned_text:
        return cleaned_text
    raise ValueError("Word 文档中未提取到可用文本，请确认文件内容不是空白。")


def _extract_spreadsheet_text(path: Path) -> str:
    document = extract_spreadsheet_document(path)
    cleaned_text = _cleanup_extracted_text(document.to_text())
    if cleaned_text:
        return cleaned_text
    raise ValueError("Excel 表格中未提取到可用内容，请确认工作表中存在非空单元格。")


def extract_spreadsheet_document(path: Path) -> SpreadsheetDocumentData:
    """从单个电子表格工作簿中提取结构化工作表行数据。"""

    suffix = path.suffix.lower()
    if suffix in OPENXML_SPREADSHEET_EXTENSIONS:
        worksheets = _extract_openxml_spreadsheet_worksheets(path)
    elif suffix in LEGACY_SPREADSHEET_EXTENSIONS:
        worksheets = _extract_legacy_spreadsheet_worksheets(path)
    else:
        display_suffix = suffix or "无扩展名"
        raise UnsupportedFileTypeError(f"暂不支持 {display_suffix}，目前仅支持 {SUPPORTED_EXTENSION_DISPLAY}")

    if not worksheets:
        raise ValueError("Excel 表格中未提取到可用内容，请确认工作表中存在非空单元格。")

    return SpreadsheetDocumentData(
        file_type=suffix.removeprefix("."),
        worksheets=worksheets,
    )


def _extract_openxml_spreadsheet_worksheets(path: Path) -> list[SpreadsheetWorksheetData]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl 依赖，无法导入 .xlsx/.xlsm 文件。") from exc

    try:
        workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError("Excel 文件解析失败，请确认文件未损坏且为有效的 .xlsx/.xlsm 文件。") from exc

    try:
        worksheets: list[SpreadsheetWorksheetData] = []
        for worksheet_index, worksheet in enumerate(workbook.worksheets):
            if str(getattr(worksheet, "sheet_state", "visible")).lower() != "visible":
                continue
            rows = _extract_sheet_rows(worksheet.iter_rows(values_only=True))
            if rows:
                worksheets.append(
                    SpreadsheetWorksheetData(
                        name=worksheet.title.strip() or f"Sheet{worksheet_index + 1}",
                        index=worksheet_index,
                        rows=rows,
                    )
                )
        return worksheets
    finally:
        workbook.close()


def _extract_legacy_spreadsheet_worksheets(path: Path) -> list[SpreadsheetWorksheetData]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("缺少 xlrd 依赖，无法导入 .xls 文件。") from exc

    try:
        workbook = xlrd.open_workbook(str(path))
    except Exception as exc:  # noqa: BLE001
        raise ValueError("Excel 文件解析失败，请确认文件未损坏且为有效的 .xls 文件。") from exc

    worksheets: list[SpreadsheetWorksheetData] = []
    for sheet_index, sheet in enumerate(workbook.sheets()):
        rows = _extract_xls_sheet_rows(sheet=sheet, workbook=workbook, xlrd_module=xlrd)
        if rows:
            worksheets.append(
                SpreadsheetWorksheetData(
                    name=sheet.name.strip() or f"Sheet{sheet_index + 1}",
                    index=sheet_index,
                    rows=rows,
                )
            )
    return worksheets


def _extract_sheet_rows(raw_rows: Any) -> list[list[str]]:
    rows: list[list[str]] = []
    for raw_row in raw_rows:
        normalized_row = [_normalize_cell_value(value) for value in raw_row]
        if any(normalized_row):
            rows.append(normalized_row)
    return rows


def _extract_xls_sheet_rows(*, sheet: Any, workbook: Any, xlrd_module: Any) -> list[list[str]]:
    rows: list[list[str]] = []
    for row_index in range(sheet.nrows):
        normalized_row: list[str] = []
        for column_index in range(sheet.ncols):
            cell = sheet.cell(row_index, column_index)
            cell_value: Any = cell.value
            if cell.ctype == xlrd_module.XL_CELL_DATE:
                try:
                    cell_value = xlrd_module.xldate_as_datetime(cell.value, workbook.datemode)
                except Exception:  # noqa: BLE001
                    cell_value = cell.value
            elif cell.ctype == xlrd_module.XL_CELL_BOOLEAN:
                cell_value = bool(cell.value)
            normalized_row.append(_normalize_cell_value(cell_value))
        if any(normalized_row):
            rows.append(normalized_row)
    return rows


def _format_sheet_block(sheet_name: str, rows: list[list[str]]) -> str:
    title = sheet_name.strip() or "未命名工作表"
    non_empty_rows = [row for row in rows if any(row)]
    if not non_empty_rows:
        return ""

    header_row = non_empty_rows[0]
    body_rows = non_empty_rows[1:]
    formatted_lines: list[str] = [f"工作表：{title}"]

    if body_rows and _looks_like_header_row(header_row):
        formatted_lines.append(f"表头：{_join_row_values(header_row)}")
        for row_index, row in enumerate(body_rows, start=2):
            record_parts: list[str] = []
            for column_index, header in enumerate(header_row):
                if column_index >= len(row):
                    continue
                value = row[column_index]
                if not value:
                    continue
                key = header or f"列{column_index + 1}"
                record_parts.append(f"{key}={value}")
            extra_cells = row[len(header_row):]
            record_parts.extend(value for value in extra_cells if value)
            if record_parts:
                formatted_lines.append(f"第 {row_index} 行：{'；'.join(record_parts)}")
        return "\n".join(formatted_lines)

    for row_index, row in enumerate(non_empty_rows, start=1):
        formatted_lines.append(f"第 {row_index} 行：{_join_row_values(row)}")
    return "\n".join(formatted_lines)


def _looks_like_header_row(row: list[str]) -> bool:
    non_empty_cells = [cell.strip() for cell in row if cell.strip()]
    if len(non_empty_cells) < 2:
        return False
    lowered_cells = [cell.lower() for cell in non_empty_cells]
    if len(set(lowered_cells)) != len(lowered_cells):
        return False
    return all(HEADER_VALUE_PATTERN.match(cell) for cell in non_empty_cells)


def _join_row_values(row: list[str]) -> str:
    values = [value for value in row if value]
    return " | ".join(values)


def _normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.15g}"
    if isinstance(value, Decimal):
        normalized = format(value, "f").rstrip("0").rstrip(".")
        return normalized or "0"

    text = str(value).strip()
    text = text.replace("\u00a0", " ").replace("\ufeff", "").replace("\x00", "")
    return re.sub(r"[ \t]+", " ", text)


def _text_file_candidate_encodings(raw_bytes: bytes) -> tuple[str, ...]:
    if raw_bytes.startswith(b"\xef\xbb\xbf"):
        return ("utf-8-sig", "utf-8", "gb18030", "gbk")
    if raw_bytes.startswith(b"\xff\xfe"):
        return ("utf-16", "utf-16-le", "utf-8-sig", "gb18030", "gbk")
    if raw_bytes.startswith(b"\xfe\xff"):
        return ("utf-16", "utf-16-be", "utf-8-sig", "gb18030", "gbk")

    nul_ratio = raw_bytes.count(b"\x00") / max(len(raw_bytes), 1)
    if nul_ratio >= 0.2:
        return ("utf-16-le", "utf-16-be", "utf-8-sig", "gb18030", "gbk")
    return ("utf-8-sig", "gb18030", "gbk", "utf-16", "utf-16-le", "utf-16-be")


def _cleanup_extracted_text(text: str) -> str:
    normalized_text = text.replace("\u00a0", " ").replace("\ufeff", "").replace("\x00", "")
    normalized_text = re.sub(r"\r\n?", "\n", normalized_text)
    raw_blocks = re.split(r"\n\s*\n+", normalized_text)
    cleaned_blocks: list[str] = []

    for raw_block in raw_blocks:
        lines = [
            re.sub(r"[ \t]+", " ", line).strip()
            for line in raw_block.split("\n")
            if line.strip()
        ]
        if not lines:
            continue
        cleaned_blocks.extend(_merge_wrapped_lines(lines))

    return "\n\n".join(cleaned_blocks).strip()


def _extract_docx_blocks(document: DocxDocumentType) -> list[str]:
    block_texts: list[str] = []

    for child in document.element.body.iterchildren():
        tag_name = child.tag.rsplit("}", maxsplit=1)[-1]
        if tag_name == "p":
            paragraph = Paragraph(child, document)
            paragraph_text = paragraph.text.strip()
            if paragraph_text:
                block_texts.append(paragraph_text)
            continue

        if tag_name == "tbl":
            table = Table(child, document)
            table_rows: list[str] = []
            for row in table.rows:
                row_cells: list[str] = []
                for cell in row.cells:
                    cell_text = " ".join(
                        paragraph.text.strip() for paragraph in cell.paragraphs if paragraph.text.strip()
                    ).strip()
                    if cell_text:
                        row_cells.append(cell_text)
                if row_cells:
                    table_rows.append(" | ".join(row_cells))
            if table_rows:
                block_texts.append("\n".join(table_rows))

    return block_texts


def _merge_wrapped_lines(lines: list[str]) -> list[str]:
    merged_blocks: list[str] = []
    current_line = ""

    for line in lines:
        if not current_line:
            current_line = line
            continue

        if LIST_PREFIX_PATTERN.match(line) or SENTENCE_ENDING_PATTERN.search(current_line):
            merged_blocks.append(current_line)
            current_line = line
            continue

        if current_line.endswith("-"):
            current_line = f"{current_line[:-1]}{line.lstrip()}"
            continue

        current_line = f"{current_line} {line}"

    if current_line:
        merged_blocks.append(current_line)
    return merged_blocks
